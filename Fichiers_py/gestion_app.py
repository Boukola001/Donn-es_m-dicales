#########################################################################
###### Noms Prénoms: BIAYE Bernadette - DIAGNE Ndeye Fatou - GBAYE Boukola - OTCHOFFA Karline
######
###### 
#########################################################################

#########################################################################
###### MODULE de gestion de l'application
#########################################################################
#importer les deux modules de gestionnaire de bdd et nettoyage
import Fichiers_py.gestion_nettoyage as gn
import Fichiers_py.gestion_bdd as gb

from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
import os

def inserer_feuille_et_graphe(res, nom_fichier, nom_feuille, titre, type_diag="pie", titre_x=None,titre_y=None):
    """
    Insère une feuille dans un fichier Excel avec un graphique (pie ou bar).

    Args:
        res (list[dict]): Résultat d'une requête SQL sous forme de liste de dictionnaires
                          avec deux clés (ex : 'genre_pat' et 'nb_patients')
        nom_fichier (str): Nom du fichier Excel (ex : "stats.xlsx")
        nom_feuille (str): Nom de la feuille à insérer
        titre (str): Titre du graphique
        titre_x (str): Titre de l'axe X du graphique en barres
        titre_y (str): Titre de l'axe Y du graphique en barres
        type_diag (str): Type de graphique ("pie" ou "bar")
    """
    if not res:
        print("Données vides : rien à insérer.")
        return

    # Déduire dynamiquement les clés (colonnes)
    colonnes = list(res[0].keys())

    # Charger ou créer le classeur Excel
    if os.path.exists(nom_fichier):
        wb = load_workbook(nom_fichier)
    else:
        wb = Workbook()

    #Supprimer la feuille par défaut si elle existe
    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)
    # Supprimer la feuille si elle existe déjà
    if nom_feuille in wb.sheetnames:
        del wb[nom_feuille]

    # Créer une nouvelle feuille
    ws = wb.create_sheet(title=nom_feuille)

    # Ajouter l'en-tête
    ws.append([col for col in colonnes])

    # Ajouter les lignes de données
    for ligne in res:
        ws.append([ligne[col] for col in colonnes])

    # Définir les références pour le graphique
    n = len(res)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)

    # Créer le graphique
    if type_diag == "pie":
        diag = PieChart()
        diag.title = f"{titre}"
        diag.add_data(data_ref, titles_from_data=True)
        diag.set_categories(labels_ref)
    elif type_diag == "bar":
        diag = BarChart()
        diag.title = f" {titre}"
        diag.y_axis.title = titre_y
        diag.x_axis.title = titre_x
        diag.add_data(data_ref, titles_from_data=True)
        diag.set_categories(labels_ref)
    else:
        print("Type de graphique invalide (choisir 'pie' ou 'bar').")
        return

    # Ajouter le graphique dans la feuille
    ws.add_chart(diag, "E2")

    # Sauvegarder
    wb.save(nom_fichier)



def orchestrer_app(nom_fichier, nom_bdd):
    """
    Cette fonction orchestre le fonctionnement de votre application. 
    Voici les étapes à faire:
    1- Créer une seule instance de connexion
    2- Vérifier si la bdd et les tables existent, si oui allez à l'étape 4 relative à l'exploration 
       des données. 
       Sinon, 
            a- Créer la bdd et les différentes tables
            b- Appliquer le nettoyage sur le jeu de données et récupérer 
            un jeu nettoyé
            c- Insérer le données dans les tables
    4- Explorer les données à partir de la bdd, générer les stats et sauvegarder
    dans un fichier Excel (pour les plus avancés visualiser à l'aide des diagrammes)
    5- Fermer la connexion
    
    Args:
    nom_fichier (str): le nom du fichier à lire
    nom_bdd (str): le nom de la BDD à explorer    
    """ 
    
    liste_tables = ["Docteurs", "Hopitaux", "Doct_Hosp", "Chambres", "Medicaments", "Patients",
                    "Condits", "Assurances", "Assu_Pat", "Tests", "Hospitalisations"]
    
    nom_host = "localhost"
    nom_user = "root"
    mdp_user = "toulouse"
    connexion = gb.creer_connexion(nom_host, nom_user, mdp_user)
    
    if not gb.bdd_existe(connexion, nom_bdd) :
        print(f"La base de données {nom_bdd} n'existe pas, on la crée.")
        gb.creer_bdd(connexion, nom_bdd)
    gb.selectionner_bdd(connexion, nom_bdd)
    if not gb.tables_deja_creees(connexion, liste_tables):
        gb.creer_tables(connexion)
        df = gn.nettoyer_preparer_donnees(nom_fichier)
        print(df)
        gb.inserer_donnees(connexion, df)
    
    
    condition = "genre"
    annee = 2021
    pat_par = gb.get_nbre_patients_par(connexion, condition)
    inserer_feuille_et_graphe(pat_par, "stats.xlsx", f"Nombre de patient par {condition}", 
                              f"Répartition des patients selon le {condition}", type_diag="pie", titre_x=None,titre_y=None)
    
    top_doc = gb.get_top10_medecins_les_plus_consultes(connexion)
    inserer_feuille_et_graphe(top_doc, "stats.xlsx", "Top10 médécins", "Top 10 des Médécins les plus consultés", 
                              type_diag="bar", titre_x="Médécins",titre_y="Nombre de consultations")
    
    top_hosp = gb.get_top5_hopitaux_a_plus_gros_budget(connexion)
    inserer_feuille_et_graphe(top_hosp, "stats.xlsx", "Top5_Hopitaux", "Top 5 des Hôpitaux par Budget", 
                              type_diag="bar", titre_x="Hopitaux",titre_y="Budget")
    
    sej_an = gb.get_nbre_sejours_annee(connexion)
    inserer_feuille_et_graphe(sej_an, "stats.xlsx", "Séjours par années", "Répartition des séjours par année", 
                              type_diag="bar", titre_x="Année",titre_y="Nombre de séjours hospitaliers")
    sej_mois = gb.get_nbre_sejours_par_mois(connexion, annee)
    inserer_feuille_et_graphe(sej_mois, "stats.xlsx", "nbre_sejours_par_mois", "Répartition mensuelle des séjours hospitaliers en {annee}", 
                              type_diag="bar", titre_x="mois",titre_y="effectifs")
 
    

    
    gb.fermer_connexion(connexion)

if __name__ == "__main__":
    nom_fichier = nom_fichier = "data/jeu_donnees_sante.csv"
    nom_bdd = "bdd_soins_sante"
    
    #appel de la fonction 
    orchestrer_app(nom_fichier, nom_bdd)
    #Partie exploration et reporting
    
    